#!/usr/bin/env python3
"""
lm_reglas.py - Modulo central de logica para el Comprobador de LM.

Usado tanto por lm_checker.py (linea de comandos) como por
lm_checker_gui.py (interfaz grafica). No se ejecuta directamente.

Todo lo que hace este modulo es DETECTAR Y AVISAR. Nunca modifica el
LM ni el catalogo master.
"""
import re
import zipfile
import pandas as pd
from lxml import etree
from odf.opendocument import load
from odf.table import Table, TableRow, TableCell
from odf import teletype
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

MAX_SALTOS = 25
RE_SUBCOMPONENTE = re.compile(r'^EN-\d+-\d+$')
HUECOS_DUROS = {'Falta P/N', 'Falta cantidad', 'Falta Gr', 'Falta nombre componente'}
NOMBRE_HOJA_CATALOGO = "Data          "  # nombre exacto de la hoja del catalogo

# ============================================================
# REGLAS ESPECIALES DEL DEPARTAMENTO
# Editar aqui si cambian las normas. No tocar el resto del fichero
# 
# ============================================================

# A) Sustituciones fijas: esta lista manda SIEMPRE por encima del catalogo
# master (nunca se modifica el master, solo se usa este valor para avisar).
OVERRIDE_OBSOLETOS = {
    'EN-0208.1': 'EN-0208.2',
    'EN-2155': 'EN-2155.2',
    'EN-1824': 'EN-2098',   # Licencia Warpalizer -> EN-2098 Warpalizer Client 1ch black level
    'EN-2270.1': 'EN-2270',
    'EN-1350': 'EN-3842',
    'EN-1968': 'EN-4002',
    'EN-2871': 'EN-2871.1',
}

# B) Correcciones de grupo: si el Gr actual no coincide con el valor correcto,
# se avisa (independientemente de cual sea el valor actual).
CORRECCIONES_GRUPO = {
    'EN-3149': '1',
    'EN-3172': '1',
    'EN-3173': '1',
    'EN-3765': 'BULK',
    'EN-3144': 'BULK',
    'EN-3316': '1',  # Night filter IPREM: de G2 a G1
}

# C) Listas usadas en las reglas condicionadas (sin checkbox, siempre se comprueban)
PB_LUM = ['EN-3159', 'EN-3156', 'EN-3240', 'EN-3157', 'EN-3158', 'EN-3155']
REF_TAPITA_PB = 'EN-4408'

PCB_REQUIERE_CPU_ETH = ['PCB200.1', 'PCB129.1', 'PCB130.1', 'PCB009.1', 'EN-1902', 'PCB235.2', 'PCB080']
REF_CPU_ETH = 'EN-2629'

REF_PCB239 = 'PCB239'
REF_PLATE_POE = 'EN-4610'

EXTINTOR_REAL = ['EN-2841', 'EN-3005', 'EN-2362']
EXTINTOR_3D = ['EN-4173', 'EN-4174', 'EN-4175']

REF_IPAD_OBSOLETO = 'EN-2800'
IPAD_NUEVO = ['EN-5558', 'EN-5559', 'EN-4464']  # ipad, funda, adaptador

# D) Reglas con checkbox
REF_DEBRIEFING_MESA = 'EN-4249'        # mesa de debriefing
REF_DEBRIEFING_CABLE_10M = 'EN-1268'   # cable LAN 10m

FLIR_REFS = ['EN-0976', 'EN-0244', 'EN-0464', 'EN-0260', 'EN-0291', 'EN-0295', 'EN-0494', 'EN-0259']

BULK_1000X_REFS = ['EN-3020', 'EN-3022', 'EN-3038', 'EN-3025', 'EN-0496', 'EN-1091', 'EN-3037']

REF_USA_ADAPTADOR = 'EN-2641'

GRUPO_3D_ENTROL = '10'  # Si Notes contiene "3D" (en cualquier parte), el material debe ser de este grupo
PATRON_3D_ENTROL = re.compile(r'\b3D\b', re.I)

# ============================================================
# Catalogo master (.ods)
# ============================================================

def normalizar_pn(pn) -> str:
    pn = str(pn).upper().strip()
    return re.sub(r'^EN\.', 'EN-', pn)

def cargar_catalogo(ruta_ods: str, nombre_hoja: str = NOMBRE_HOJA_CATALOGO) -> dict:
    """Lee el catalogo master (.ods) y devuelve {referencia_normalizada: descripcion}.

    Implementacion con lxml.etree.iterparse en streaming directamente sobre el
    XML interno del .ods, en vez de pandas+odfpy (que construye un arbol de
    objetos completo en memoria por cada celda y tardaba ~12s con un catalogo
    de 6000+ filas). Esta version tarda ~0.5s con el mismo archivo: no
    construye ningun arbol intermedio, solo extrae las dos columnas que
    necesitamos (Referencia, Descripcion), localizadas dinamicamente por su
    nombre en la fila de cabecera (igual que hacia pandas, para que siga
    funcionando si alguien selecciona un .ods con esas columnas en otro orden).

    IMPORTANTE: cada fila se itera SIEMPRE hasta el final antes de llamar a
    elem.clear() - nunca se interrumpe un generador a medias y se limpia el
    elemento despues, porque eso corrompe la memoria interna de lxml de forma
    intermitente (provoca cuelgues aleatorios dificiles de reproducir).
    """
    NS_TABLE = 'urn:oasis:names:tc:opendocument:xmlns:table:1.0'
    NS_TEXT = 'urn:oasis:names:tc:opendocument:xmlns:text:1.0'
    T_TABLE, T_ROW = f"{{{NS_TABLE}}}table", f"{{{NS_TABLE}}}table-row"
    T_CELL, T_COVERED = f"{{{NS_TABLE}}}table-cell", f"{{{NS_TABLE}}}covered-table-cell"
    T_NAME, T_REPEAT = f"{{{NS_TABLE}}}name", f"{{{NS_TABLE}}}number-columns-repeated"
    T_S, T_C, T_TAB = f"{{{NS_TEXT}}}s", f"{{{NS_TEXT}}}c", f"{{{NS_TEXT}}}tab"

    def extraer_texto(elem) -> str:
        """Texto completo de una celda, incluyendo spans anidados y los
        espacios multiples de ODF (<text:s text:c='N'/>), que si no se tratan
        de forma especial se pierden (no tienen .text propio)."""
        partes = [elem.text] if elem.text else []
        for hijo in elem:
            if hijo.tag == T_S:
                partes.append(' ' * int(hijo.get(T_C, '1')))
            elif hijo.tag == T_TAB:
                partes.append('\t')
            else:
                partes.append(extraer_texto(hijo))
            if hijo.tail:
                partes.append(hijo.tail)
        return ''.join(partes)

    def valores_de_fila(fila_elem, columnas_buscadas):
        """Devuelve {indice_columna: texto} solo para las columnas en
        columnas_buscadas, recorriendo la fila COMPLETA (sin cortar a medias)
        para poder limpiar el elemento de forma segura despues."""
        encontrados = {}
        col = 0
        objetivo = set(columnas_buscadas)
        for celda in fila_elem:
            if celda.tag == T_CELL:
                rep = int(celda.get(T_REPEAT, '1'))
                if col in objetivo:
                    encontrados[col] = extraer_texto(celda)
                col += rep
            elif celda.tag == T_COVERED:
                col += int(celda.get(T_REPEAT, '1'))
        return encontrados

    def todas_las_celdas(fila_elem):
        """Para la fila de cabecera: {indice_columna: texto} de TODAS las celdas."""
        encontrados = {}
        col = 0
        for celda in fila_elem:
            if celda.tag == T_CELL:
                rep = int(celda.get(T_REPEAT, '1'))
                encontrados[col] = extraer_texto(celda)
                col += rep
            elif celda.tag == T_COVERED:
                col += int(celda.get(T_REPEAT, '1'))
        return encontrados

    catalogo = {}
    col_ref = col_desc = None

    with zipfile.ZipFile(ruta_ods) as z, z.open('content.xml') as f:
        contexto = etree.iterparse(f, events=('start', 'end'), tag=(T_TABLE, T_ROW))
        dentro_hoja = False
        indice_fila = 0
        for evento, elem in contexto:
            if elem.tag == T_TABLE:
                if evento == 'start':
                    dentro_hoja = (elem.get(T_NAME) == nombre_hoja)
                    indice_fila = 0
                elif dentro_hoja:
                    break  # ya hemos terminado de leer la hoja que nos interesa
                continue

            if evento != 'end' or not dentro_hoja:
                # ignoramos el evento 'start' de la fila (los hijos/celdas aun
                # no estan disponibles en ese momento) y cualquier fila fuera
                # de la hoja que nos interesa
                continue

            if indice_fila == 0:
                # fila de cabecera: localizar en que columna estan "Referencia" y "Descripción"
                valores = todas_las_celdas(elem)
                for col, texto in valores.items():
                    if texto == 'Referencia':
                        col_ref = col
                    elif texto == 'Descripción':
                        col_desc = col
                if col_ref is None or col_desc is None:
                    raise ValueError(
                        f"No se encontraron las columnas 'Referencia'/'Descripción' "
                        f"en la cabecera de la hoja '{nombre_hoja}' del catálogo."
                    )
                indice_fila += 1
                continue

            indice_fila += 1
            valores = valores_de_fila(elem, (col_ref, col_desc))
            ref_val = valores.get(col_ref)
            desc_val = valores.get(col_desc)

            if ref_val:
                ref_norm = normalizar_pn(ref_val)
                if ref_norm:
                    catalogo[ref_norm] = (desc_val or '').strip()

    if col_ref is None:
        raise ValueError(f"No se encontró la hoja '{nombre_hoja}' en el catálogo, o estaba vacía.")

    return catalogo

def aumentar_version(pn: str) -> str:
    partes = pn.split('.')
    if len(partes) == 1:
        return pn + ".1"
    return partes[0] + "." + str(int(partes[1]) + 1)

def rastrear_pn(pn_actual, catalogo, saltos=0, visitados=None):
    if visitados is None:
        visitados = []
    pn_actual = normalizar_pn(pn_actual)
    if saltos > MAX_SALTOS:
        return [f"__SIN_RESOLVER__:{pn_actual}"]
    if pn_actual in visitados:
        return rastrear_pn(aumentar_version(pn_actual), catalogo, saltos + 1, visitados)
    nuevos_visitados = visitados + [pn_actual]
    descripcion = catalogo.get(pn_actual)
    if not descripcion:
        if '.' in pn_actual:
            return rastrear_pn(aumentar_version(pn_actual), catalogo, saltos + 1, nuevos_visitados)
        return [f"__SIN_RESOLVER__:{pn_actual}"]
    if descripcion.upper().startswith("OBSOLETE"):
        pns_nuevos = []
        for frase in re.split(r'[.;\n]', descripcion):
            m = re.search(r'(?:ver\b|->|sustituid[oa]|reemplazad[oa]|nuevo\b|new\b|cambiar\b|usar\b)\s*:?\s*(.*)', frase, re.I)
            if m:
                pns_nuevos += re.findall(r'EN[-.]\d+(?:\.\d+)?', m.group(1), re.I)
        if not pns_nuevos:
            return rastrear_pn(aumentar_version(pn_actual), catalogo, saltos + 1, nuevos_visitados)
        resultado = []
        for nuevo in pns_nuevos:
            nuevo = normalizar_pn(nuevo)
            if nuevo == pn_actual:
                resultado += rastrear_pn(aumentar_version(pn_actual), catalogo, saltos + 1, nuevos_visitados)
            else:
                resultado += rastrear_pn(nuevo, catalogo, saltos + 1, nuevos_visitados)
        return resultado
    return [pn_actual]

def estado_pn(pn, catalogo):
    """Devuelve (estado, detalle). estado in {OK, OBSOLETO, NO_ENCONTRADO}.
    La lista OVERRIDE_OBSOLETOS manda SIEMPRE antes que el catalogo master."""
    pn_norm = normalizar_pn(pn)

    if pn_norm in OVERRIDE_OBSOLETOS:
        return "OBSOLETO", OVERRIDE_OBSOLETOS[pn_norm]

    descripcion = catalogo.get(pn_norm)
    if descripcion is None:
        return "NO_ENCONTRADO", None

    if descripcion.upper().startswith("OBSOLETE"):
        finales = list(dict.fromkeys(rastrear_pn(pn_norm, catalogo)))
        resueltos = [f for f in finales if not f.startswith("__SIN_RESOLVER__")]
        if resueltos:
            return "OBSOLETO", ", ".join(resueltos)
        return "NO_ENCONTRADO", None

    return "OK", None

# ============================================================
# LM (.odt)
# ============================================================

def extraer_tabla_odt(ruta_odt: str) -> pd.DataFrame:
    doc = load(ruta_odt)
    tablas = doc.getElementsByType(Table)
    if not tablas:
        raise ValueError("El documento no contiene ninguna tabla.")
    table = tablas[0]
    filas = table.getElementsByType(TableRow)

    def texto_fila(r):
        celdas = r.getElementsByType(TableCell)
        out = []
        for c in celdas:
            txt = teletype.extractText(c)
            rep = int(c.getAttribute("numbercolumnsrepeated") or 1)
            out.extend([txt] * rep)
        return out

    datos = [texto_fila(r) for r in filas]
    if len(datos) < 3:
        raise ValueError("La tabla del LM parece estar vacía o incompleta.")
    cabecera = datos[1]
    cuerpo = [r for r in datos[2:] if len(r) >= 6]
    if not cuerpo:
        raise ValueError("No se encontraron filas de datos en la tabla del LM.")
    df = pd.DataFrame(cuerpo, columns=cabecera[:6])
    df.columns = ['MTS', 'Description', 'PN', 'Qty', 'Notes', 'Gr']
    return df

def clasificar_pn(valor) -> str:
    v = str(valor).strip()
    if v in ('', 'nan'):
        return 'VACIO'
    if 'xxxx' in v.lower():
        return 'PLACEHOLDER'
    if v == '-':
        return 'DASH'
    if re.match(r'^EN-?\d+', v):
        return 'REAL_EN'
    return 'OTRO_CODIGO'

def vacio_o_dash(v) -> bool:
    return str(v).strip() in ('', '-', 'nan')

def es_candidato_pieza(fila) -> bool:
    """True si PN, Qty o Gr tiene CUALQUIER contenido.
    Si los tres estan vacios/guion, es una cabecera de sección (Notes se
    ignora a propósito: muchas cabeceras reales llevan texto explicativo
    en Notes sin dejar de ser cabeceras)."""
    return not (vacio_o_dash(fila['PN']) and vacio_o_dash(fila['Qty']) and vacio_o_dash(fila['Gr']))

def cantidad_int(valor):
    try:
        return int(str(valor).strip())
    except (ValueError, TypeError):
        return None

def prioridad_estado(tipos):
    if any(t.startswith('Obsoleto') for t in tipos):
        return 'OBSOLETO'
    if 'No encontrado en el master' in tipos:
        return 'NO_ENCONTRADO'
    if any(t in HUECOS_DUROS for t in tipos):
        return 'HUECO'
    return 'REVISAR'

# ============================================================
# Analisis fila a fila (huecos, obsoletos, no encontrados)
# ============================================================

def analizar_lm(df_lm: pd.DataFrame, catalogo: dict, progreso_cb=None):
    incidencias = []
    estados, detalles = [], []
    total = len(df_lm)

    for idx, (i, fila) in enumerate(df_lm.iterrows()):
        if progreso_cb and idx % 50 == 0:
            progreso_cb(idx, total)

        if not es_candidato_pieza(fila):
            # PN, Qty y Gr vacios/guion -> cabecera de seccion, se omite sin aviso
            estados.append('')
            detalles.append('')
            continue

        problemas = []
        desc = str(fila['Description']).strip()
        if desc in ('', 'nan'):
            problemas.append(('Falta nombre componente', 'Columna Description vacía'))

        pn_raw = fila['PN']
        pn_clase = clasificar_pn(pn_raw)

        if pn_clase in ('VACIO', 'PLACEHOLDER'):
            problemas.append(('Falta P/N', f"P/N sin rellenar ({pn_raw!r})"))
        elif pn_clase == 'DASH':
            problemas.append(('Revisar P/N', 'P/N marcado como "-": confirmar si es intencional o un olvido'))
        elif pn_clase == 'REAL_EN' and not RE_SUBCOMPONENTE.match(str(pn_raw).strip()):
            estado, detalle = estado_pn(pn_raw, catalogo)
            if estado == 'OBSOLETO':
                problemas.append(('Obsoleto', f"Reemplazo: {detalle}"))
            elif estado == 'NO_ENCONTRADO':
                problemas.append(('No encontrado en el master', 'PN no está en el catálogo master'))

        qty_raw = str(fila['Qty']).strip()
        if qty_raw in ('', 'nan'):
            problemas.append(('Falta cantidad', 'Columna Qty vacía'))
        elif qty_raw == '-':
            problemas.append(('Revisar cantidad', 'Qty marcada como "-": confirmar si es intencional o un olvido'))

        gr_raw = str(fila['Gr']).strip()
        if gr_raw in ('', '-', 'nan'):
            problemas.append(('Falta Gr', 'Columna Gr vacía o "-" en una fila que parece ser pieza real'))

        # Correccion de grupo fija (regla B), independiente de si ya hay otros problemas
        pn_norm = normalizar_pn(pn_raw)
        if pn_norm in CORRECCIONES_GRUPO:
            objetivo = CORRECCIONES_GRUPO[pn_norm]
            actual = gr_raw.upper()
            if actual != objetivo:
                problemas.append(('Revisar Gr incorrecto',
                                   f"{pn_norm}: Gr actual='{gr_raw}', debería ser '{objetivo}'"))

        if problemas:
            tipos = [p[0] for p in problemas]
            detalle_txt = "; ".join(p[1] for p in problemas)
            estados.append(prioridad_estado(tipos))
            detalles.append(detalle_txt)
            incidencias.append({
                'Fila': i + 4, 'Description': fila['Description'], 'PN': fila['PN'],
                'Qty': fila['Qty'], 'Tipo': ", ".join(tipos), 'Detalle': detalle_txt
            })
        else:
            estados.append('OK')
            detalles.append('')

    if progreso_cb:
        progreso_cb(total, total)

    df_lm = df_lm.copy()
    df_lm['Estado'] = estados
    df_lm['Detalle'] = detalles
    return df_lm, pd.DataFrame(incidencias)

# ============================================================
# Reglas especiales / condicionadas (no van fila a fila: miran todo el LM)
# ============================================================

def _filas_con_pn(df_lm, lista_pns_normalizados):
    objetivo = {normalizar_pn(p) for p in lista_pns_normalizados}
    return df_lm[df_lm['PN'].apply(normalizar_pn).isin(objetivo)]

def _existe_pn(df_lm, pn) -> bool:
    objetivo = normalizar_pn(pn)
    return (df_lm['PN'].apply(normalizar_pn) == objetivo).any()

def aplicar_reglas_especiales(df_lm: pd.DataFrame, checkboxes: dict):
    """
    checkboxes esperado:
        {'pais': 'UE' | 'USA', 'debriefing': bool, 'flir': bool, 'sim_1000x': bool}

    Devuelve una lista de incidencias (dicts) con 'Fila' = numero de fila real
    o 'General' si no esta atada a una fila concreta del LM.
    """
    avisos = []

    def agregar(fila_num, tipo, detalle):
        desc = pn = qty = gr = ''
        if fila_num != 'General':
            idx = fila_num - 4  # mismo desplazamiento usado al generar 'Fila' (i + 4)
            if idx in df_lm.index:
                fila_real = df_lm.loc[idx]
                desc, pn, qty, gr = fila_real['Description'], fila_real['PN'], fila_real['Qty'], fila_real['Gr']
        avisos.append({'Fila': fila_num, 'Description': desc, 'PN': pn, 'Qty': qty, 'Gr': gr,
                        'Tipo': tipo, 'Detalle': detalle})

    # --- pb lum -> EN-4408 (tapita), 1 unidad por cada unidad de pb lum ---
    filas_pb = _filas_con_pn(df_lm, PB_LUM)
    total_necesario = sum((cantidad_int(q) or 0) for q in filas_pb['Qty'])
    if total_necesario > 0:
        filas_tapita = _filas_con_pn(df_lm, [REF_TAPITA_PB])
        actual = sum((cantidad_int(q) or 0) for q in filas_tapita['Qty'])
        if filas_tapita.empty:
            agregar('General', f'Falta {REF_TAPITA_PB} (tapita pb)',
                     f"Hay pb lum en el LM (total {total_necesario} ud). Falta añadir {REF_TAPITA_PB} "
                     f"x{total_necesario} ud, Gr 1.")
        elif actual < total_necesario:
            agregar('General', f'Cantidad insuficiente de {REF_TAPITA_PB}',
                     f"Hay {actual} ud de {REF_TAPITA_PB}, se necesitan {total_necesario} "
                     f"(1 por cada ud de pb lum).")

    # --- PCB / EN-1902 -> EN-2629 (CPU ETH) ---
    if not _filas_con_pn(df_lm, PCB_REQUIERE_CPU_ETH).empty:
        filas_cpu = _filas_con_pn(df_lm, [REF_CPU_ETH])
        if filas_cpu.empty:
            agregar('General', f'Falta {REF_CPU_ETH} (CPU ETH)',
                     f"Hay PCBs/EN que requieren {REF_CPU_ETH}: falta añadirlo, 1 ud, Gr 1.")
        else:
            for _, f in filas_cpu.iterrows():
                if (cantidad_int(f['Qty']) or 0) < 1 or str(f['Gr']).strip().upper() != '1':
                    agregar('General', f'Revisar {REF_CPU_ETH}',
                             f"{REF_CPU_ETH} presente pero con Qty='{f['Qty']}' / Gr='{f['Gr']}' "
                             f"(debería ser 1 ud, Gr 1).")

    # --- PCB239 -> EN-4610 (plate 3D soporte POE) ---
    if _existe_pn(df_lm, REF_PCB239):
        if _filas_con_pn(df_lm, [REF_PLATE_POE]).empty:
            agregar('General', f'Falta {REF_PLATE_POE} (plate 3D soporte POE)',
                     f"Hay {REF_PCB239} en el LM: falta añadir {REF_PLATE_POE}, 1 ud, Gr 1.")

    # --- Extintor real -> 3D (EN-4173+4174+4175) ---
    if not _filas_con_pn(df_lm, EXTINTOR_REAL).empty:
        for ref in EXTINTOR_3D:
            filas_ref = _filas_con_pn(df_lm, [ref])
            if filas_ref.empty:
                agregar('General', f'Falta {ref} (extintor 3D)',
                         f"Hay extintor real en el LM: debe simularse en 3D. Falta {ref}, 1 ud, Gr 1.")
            else:
                for _, f in filas_ref.iterrows():
                    if (cantidad_int(f['Qty']) or 0) < 1 or str(f['Gr']).strip().upper() != '1':
                        agregar('General', f'Revisar {ref}',
                                 f"{ref} presente pero con Qty='{f['Qty']}' / Gr='{f['Gr']}' (debería ser 1 ud, Gr 1).")

    # --- EN-2800 (IPAD) obsoleto -> EN-2800.2 + funda + adaptador ---
    if _existe_pn(df_lm, REF_IPAD_OBSOLETO):
        for ref in IPAD_NUEVO:
            if _filas_con_pn(df_lm, [ref]).empty:
                agregar('General', f'Falta {ref}',
                         f"{REF_IPAD_OBSOLETO} es obsoleto: debe pasar a EN-2800.2 + funda EN-4430 + "
                         f"adaptador EN-4464. Falta {ref}, 1 ud, Gr 1.")

    # --- Checkbox: Debriefing -> añadir mesa (EN-4249) y cable LAN 10m (EN-1268) ---
    if checkboxes.get('debriefing'):
        for ref, nombre in [(REF_DEBRIEFING_MESA, 'mesa de debriefing'),
                             (REF_DEBRIEFING_CABLE_10M, 'cable LAN 10m')]:
            if _filas_con_pn(df_lm, [ref]).empty:
                agregar('General', f'Falta {ref} ({nombre})',
                         f"Se ha marcado Debriefing: falta añadir {ref} ({nombre}), 1 ud.")

    # --- Checkbox: FLIR handheld -> quitar referencias asociadas ---
    if checkboxes.get('flir'):
        for ref in FLIR_REFS:
            if _existe_pn(df_lm, ref):
                agregar('General', f'Quitar {ref} (FLIR handheld)',
                         f"Se ha marcado FLIR handheld: {ref} debería eliminarse del LM.")

    # --- Checkbox: Simulador 1000x -> esas refs deben ser Gr BULK ---
    if checkboxes.get('sim_1000x'):
        for ref in BULK_1000X_REFS:
            for _, f in _filas_con_pn(df_lm, [ref]).iterrows():
                if str(f['Gr']).strip().upper() != 'BULK':
                    agregar('General', f'Revisar Gr de {ref} (Simulador 1000x)',
                             f"{ref}: Gr actual='{f['Gr']}', debería ser 'BULK' por ser simulador 1000x.")

    # --- Notes contiene "3D Entrol" -> el grupo debe ser 10 (sin checkbox, siempre se comprueba) ---
    for i, fila in df_lm.iterrows():
        notes = str(fila.get('Notes', ''))
        if PATRON_3D_ENTROL.search(notes):
            gr_actual = str(fila['Gr']).strip()
            if gr_actual.upper() != GRUPO_3D_ENTROL:
                agregar(i + 4, 'Revisar Gr (3D Entrol)',
                        f"Notes indica '3D Entrol': Gr actual='{fila['Gr']}', debería ser '{GRUPO_3D_ENTROL}'.")

    # --- Checkbox: USA -> EN-2641 requerido + aviso por tensiones 230V ---
    if checkboxes.get('pais') == 'USA':
        if not _existe_pn(df_lm, REF_USA_ADAPTADOR):
            agregar('General', f'Falta {REF_USA_ADAPTADOR} (adaptador USA)',
                     f"Destino USA: falta {REF_USA_ADAPTADOR} (adaptador para tooling de módulos de pantalla LED).")

        patron_230v = re.compile(r'230\s*v\b', re.I)
        for i, fila in df_lm.iterrows():
            desc = str(fila['Description'])
            if patron_230v.search(desc):
                agregar(i + 4, 'Revisar tensión (destino USA)',
                         f"Descripción menciona 230V con destino USA: '{desc.strip()}'. Confirmar si es correcto.")

    return avisos

# ============================================================
# Verificacion de ultimas referencias (PLT, PNL, MCD, LEG, KNB, KEY)
# Funcion opcional, activada por checkbox "Verificar últimas referencias"
# ============================================================

PREFIJOS_REFERENCIA = ['PLT', 'PNL', 'MCD', 'LEG', 'KNB', 'KEY']

def extraer_cabecera_modelo(ruta_odt: str) -> dict:
    """Lee de la cabecera de pagina del documento (styles.xml -> style:header)
    el modelo de simulador (ej. 'H64' de 'PARTS CATALOG H64') y los valores
    declarados para cada prefijo (ej. 'PLT72, PNL59, ...').
    Devuelve {'modelo': 'H64', 'declarado': {'PLT': 72, 'PNL': 59, ...}}."""
    NS_STYLE = 'urn:oasis:names:tc:opendocument:xmlns:style:1.0'
    NS_TEXT = 'urn:oasis:names:tc:opendocument:xmlns:text:1.0'
    T_HEADER = f"{{{NS_STYLE}}}header"
    T_P = f"{{{NS_TEXT}}}p"

    with zipfile.ZipFile(ruta_odt) as z:
        styles_xml = z.read('styles.xml')
    root = etree.fromstring(styles_xml)
    header = root.find(f'.//{T_HEADER}')
    if header is None:
        return {'modelo': None, 'declarado': {}}

    parrafos = header.findall(T_P)
    texto = '\n'.join(''.join(p.itertext()) for p in parrafos)

    m_modelo = re.search(r'PARTS CATALOG\s+([A-Z]{1,3}\d+)', texto, re.I)
    modelo = m_modelo.group(1).upper() if m_modelo else None

    declarado = {}
    for prefijo in PREFIJOS_REFERENCIA:
        m = re.search(re.escape(prefijo) + r'(\d+)', texto)
        if m:
            declarado[prefijo] = int(m.group(1))

    return {'modelo': modelo, 'declarado': declarado}


def analizar_referencias_modelo(df_lm: pd.DataFrame, modelo_actual: str, declarado: dict) -> list:
    """Para cada prefijo (PLT, PNL, MCD, LEG, KNB, KEY) busca en la columna PN
    referencias con formato PREFIJO.MODELO.NUMERO (el MODELO no puede ser
    puramente numerico - eso descarta formatos invalidos tipo PNL.0263 o
    PNL178.1 sin modelo). Calcula el numero mas alto encontrado, tanto en
    general (cualquier modelo) como especificamente para el modelo actual,
    y lo compara con lo que declara la cabecera."""
    filas = []
    for prefijo in PREFIJOS_REFERENCIA:
        patron = re.compile(rf'^{prefijo}\.([A-Za-z]\w*)\.(\d+)(?:\.\d+)?$', re.I)
        encontrados = []  # (modelo, numero)
        for pn in df_lm['PN']:
            m = patron.match(str(pn).strip())
            if m:
                encontrados.append((m.group(1).upper(), int(m.group(2))))

        cabecera_dice = declarado.get(prefijo)
        avisos_celda = []

        if not encontrados:
            max_actual = max_global = modelo_global = None
            avisos_celda.append(f"No se encontró ninguna referencia {prefijo}.{modelo_actual}.NN en el LM.")
        else:
            modelo_global, max_global = max(encontrados, key=lambda x: x[1])
            del_actual = [n for m_, n in encontrados if m_ == modelo_actual]
            max_actual = max(del_actual) if del_actual else None

            if max_actual is None:
                avisos_celda.append(f"No se encontró ninguna referencia del modelo actual ({modelo_actual}); "
                                     f"la más alta encontrada es {prefijo}.{modelo_global}.{max_global}, de otro simulador.")
            elif modelo_global != modelo_actual:
                avisos_celda.append(f"El número más alto encontrado es {max_global} pero pertenece a otro "
                                     f"simulador ({modelo_global}); el más alto de este simulador "
                                     f"({modelo_actual}) es {max_actual}.")

            if cabecera_dice is not None and max_actual is not None and cabecera_dice != max_actual:
                avisos_celda.append(f"La cabecera indica {cabecera_dice} pero el máximo real de este "
                                     f"simulador es {max_actual}.")

        filas.append({
            'Prefijo': prefijo,
            'Modelo actual': modelo_actual,
            'Cabecera indica': cabecera_dice if cabecera_dice is not None else '',
            'Máximo en este simulador': max_actual if max_actual is not None else '',
            'Máximo global encontrado': max_global if max_global is not None else '',
            'Modelo del máximo global': modelo_global if modelo_global is not None else '',
            'Aviso': ' '.join(avisos_celda),
        })
    return filas


# ============================================================
# Salida
# ============================================================

COLORES_HEX = {
    'OBSOLETO': 'FFC7CE',
    'HUECO': 'FFEB9C',
    'NO_ENCONTRADO': 'FFD699',
    'REVISAR': 'D9E1F2',
}

ORDEN_PRIORIDAD = ['', 'OK', 'REVISAR', 'HUECO', 'NO_ENCONTRADO', 'OBSOLETO']

def _fusionar_avisos_en_filas(df_lm: pd.DataFrame, avisos_generales):
    """Combina los avisos atados a una fila concreta (los que generan
    aplicar_reglas_especiales con Fila != 'General') en las columnas
    Estado/Detalle de esa misma fila, para que se vean resaltados junto al
    material al que pertenecen en la hoja principal, en vez de solo en una
    lista aparte sin contexto."""
    df = df_lm.copy()
    for av in avisos_generales or []:
        fila_num = av.get('Fila')
        if fila_num == 'General' or fila_num is None:
            continue
        idx = fila_num - 4
        if idx not in df.index:
            continue
        nuevo_estado = prioridad_estado([av['Tipo']])
        estado_actual = df.at[idx, 'Estado'] or ''
        rango_actual = ORDEN_PRIORIDAD.index(estado_actual) if estado_actual in ORDEN_PRIORIDAD else 1
        rango_nuevo = ORDEN_PRIORIDAD.index(nuevo_estado)
        if rango_nuevo > rango_actual:
            df.at[idx, 'Estado'] = nuevo_estado
        detalle_actual = df.at[idx, 'Detalle'] or ''
        nuevo_detalle = f"{av['Tipo']}: {av['Detalle']}"
        df.at[idx, 'Detalle'] = f"{detalle_actual}; {nuevo_detalle}" if detalle_actual else nuevo_detalle
    return df

def generar_excel(df_lm: pd.DataFrame, ruta_xlsx: str, avisos_generales=None, tabla_referencias=None):
    df_lm = _fusionar_avisos_en_filas(df_lm, avisos_generales)

    wb = Workbook()
    ws = wb.active
    ws.title = "LM revisado"

    columnas = ['MTS', 'Description', 'PN', 'Qty', 'Notes', 'Gr', 'Estado', 'Detalle']
    ws.append(columnas)
    for c in range(1, len(columnas) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    for _, fila in df_lm.iterrows():
        ws.append([fila[c] for c in columnas])

    for r in range(2, ws.max_row + 1):
        estado = ws.cell(row=r, column=7).value
        if estado in COLORES_HEX:
            fill = PatternFill('solid', start_color=COLORES_HEX[estado], end_color=COLORES_HEX[estado])
            for c in range(1, len(columnas) + 1):
                ws.cell(row=r, column=c).fill = fill

    anchos = [30, 45, 22, 8, 35, 6, 22, 60]
    for i, w in enumerate(anchos, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Hoja aparte con TODOS los avisos de reglas especiales (tanto los atados
    # a una fila como los generales), con Fila/Description/PN/Gr para poder
    # localizar el material aunque no se haya mirado la hoja principal.
    if avisos_generales:
        ws2 = wb.create_sheet("Avisos generales")
        cols2 = ['Fila', 'Description', 'PN', 'Gr', 'Tipo', 'Detalle']
        ws2.append(cols2)
        for c in range(1, len(cols2) + 1):
            ws2.cell(row=1, column=c).font = Font(bold=True)
        fill_aviso = PatternFill('solid', start_color=COLORES_HEX['REVISAR'], end_color=COLORES_HEX['REVISAR'])
        for av in avisos_generales:
            ws2.append([av.get('Fila', 'General'), av.get('Description', ''), av.get('PN', ''),
                        av.get('Gr', ''), av['Tipo'], av['Detalle']])
            for c in range(1, len(cols2) + 1):
                ws2.cell(row=ws2.max_row, column=c).fill = fill_aviso
        anchos2 = [8, 40, 18, 8, 30, 70]
        for i, w in enumerate(anchos2, start=1):
            ws2.column_dimensions[get_column_letter(i)].width = w

    # Hoja aparte con la verificacion de ultimas referencias (PLT/PNL/MCD/LEG/KNB/KEY),
    # solo si se ha pedido (checkbox "Verificar últimas referencias")
    if tabla_referencias:
        ws3 = wb.create_sheet("Últimas referencias")
        cols3 = ['Prefijo', 'Modelo actual', 'Cabecera indica', 'Máximo en este simulador',
                 'Máximo global encontrado', 'Modelo del máximo global', 'Aviso']
        ws3.append(cols3)
        for c in range(1, len(cols3) + 1):
            ws3.cell(row=1, column=c).font = Font(bold=True)
        fill_aviso_ref = PatternFill('solid', start_color=COLORES_HEX['REVISAR'], end_color=COLORES_HEX['REVISAR'])
        for fila_ref in tabla_referencias:
            ws3.append([fila_ref[c] for c in cols3])
            if fila_ref['Aviso']:
                for c in range(1, len(cols3) + 1):
                    ws3.cell(row=ws3.max_row, column=c).fill = fill_aviso_ref
        anchos3 = [10, 14, 16, 22, 22, 22, 80]
        for i, w in enumerate(anchos3, start=1):
            ws3.column_dimensions[get_column_letter(i)].width = w

    wb.save(ruta_xlsx)
